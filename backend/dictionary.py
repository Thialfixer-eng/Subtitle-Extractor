import os
import re
import fnmatch
import urllib.request
import gzip

DICT_URL = "https://www.mdbg.net/chinese/export/cedict/cedict_1_0_ts_utf-8_mdbg.txt.gz"
PLECO_URL = "https://github.com/jimmy-zhening-luo/pleco-mega-big-chinese-dictionary/raw/master/big-table.xlsx"
WIKDICT_PL_URL = "https://download.wikdict.com/dictionaries/sqlite/2_2026-06/zh-pl.sqlite3"
CACHE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "cache")
CACHE_FILE = os.path.join(CACHE_DIR, "cedict.txt")
PLECO_CACHE = os.path.join(CACHE_DIR, "big-table.xlsx")
WIKDICT_PL_CACHE = os.path.join(CACHE_DIR, "zh-pl_wikdict.sqlite3")

ENTRY_RE = re.compile(r"^(\S+) (\S+) \[(.+?)\] /(.+)/$")

# Chinese Unicode ranges
CJK_RANGE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")
# Pinyin chars (a-z with tone number or diacritic)
PY_RANGE = re.compile(r"^[a-zA-Z\u0100-\u024F0-9 ]+$")
# Tone numbers
TONE_RE = re.compile(r"[0-9]")


def _ensure_dict():
    if os.path.exists(CACHE_FILE):
        return
    os.makedirs(CACHE_DIR, exist_ok=True)
    print("[DICT] Downloading CC-CEDICT...")
    urllib.request.urlretrieve(DICT_URL, CACHE_FILE + ".gz")
    with gzip.open(CACHE_FILE + ".gz", "rt", encoding="utf-8") as f:
        content = f.read()
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        f.write(content)
    os.remove(CACHE_FILE + ".gz")
    print(f"[DICT] Saved to {CACHE_FILE}")


def load_entries():
    _ensure_dict()
    entries = []
    with open(CACHE_FILE, "r", encoding="utf-8") as f:
        for line in f:
            if line.startswith("#") or line.startswith("%"):
                continue
            m = ENTRY_RE.match(line.strip())
            if m:
                trad, simp, pinyin, defs = m.groups()
                entries.append({
                    "trad": trad,
                    "simp": simp,
                    "pinyin": pinyin,
                    "defs": defs.split("/"),
                })
    return entries


def build_index(entries):
    idx = {}
    for e in entries:
        for key in (e["simp"], e["trad"]):
            idx.setdefault(key, []).append(e)
    return idx


def _norm_pinyin(py):
    """Normalize pinyin for matching: remove tone numbers, collapse spaces, lowercase."""
    py = TONE_RE.sub("", py).strip().lower()
    py = py.replace("'", "").replace("  ", " ")
    # CEDICT uses u: for ü; also accept v as input
    py = py.replace("u:", "\u00fc").replace("v", "\u00fc")
    return py


# --- Pleco metadata ---

def _ensure_pleco():
    if os.path.exists(PLECO_CACHE):
        return
    os.makedirs(CACHE_DIR, exist_ok=True)
    print("[DICT] Downloading Pleco Mega Dictionary...")
    urllib.request.urlretrieve(PLECO_URL, PLECO_CACHE)
    print(f"[DICT] Saved to {PLECO_CACHE}")


def _load_pleco_export():
    _ensure_pleco()
    try:
        import openpyxl
    except ImportError:
        print("[DICT] openpyxl not installed, skipping Pleco metadata")
        return {}
    wb = openpyxl.load_workbook(PLECO_CACHE, data_only=True, read_only=True)
    ws = wb["PlecoExport"]
    meta = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[9] is None:
            continue
        simp = str(row[9]).strip()
        meta[simp] = {
            "trad": str(row[10] or "").strip(),
            "pinyin": str(row[1] or "").strip(),
            "hsk": str(row[8] or "").strip(),
            "strokes": str(row[6] or "").strip(),
            "giga_rank": str(row[2] or "").strip(),
            "junda_rank": str(row[3] or "").strip(),
        }
    wb.close()
    return meta


# --- WikDict Chinese-Polish ---

def _ensure_wikdict_pl():
    if os.path.exists(WIKDICT_PL_CACHE):
        return
    os.makedirs(CACHE_DIR, exist_ok=True)
    print("[DICT] Downloading Chinese-Polish dictionary (WikDict)...")
    urllib.request.urlretrieve(WIKDICT_PL_URL, WIKDICT_PL_CACHE)
    print(f"[DICT] Saved to {WIKDICT_PL_CACHE}")


def _load_wikdict_pl():
    _ensure_wikdict_pl()
    try:
        import sqlite3
    except ImportError:
        print("[DICT] sqlite3 not available, skipping Polish dictionary")
        return [], {}
    conn = sqlite3.connect(WIKDICT_PL_CACHE)
    c = conn.cursor()
    entries = []
    for row in c.execute("SELECT written_rep, trans_list FROM simple_translation"):
        zh = row[0].strip()
        pl = row[1]
        if not zh or not pl:
            continue
        # strip /翻譯 suffix (compound phrasal entries in Wiktionary)
        zh = re.sub(r"/翻譯$", "", zh).strip()
        defs = [d.strip() for d in pl.split("|")]
        entries.append({
            "trad": zh,
            "simp": zh,
            "pinyin": "",
            "defs": defs,
            "hsk": "",
            "strokes": "",
        })
    conn.close()
    idx = {}
    for e in entries:
        for key in (e["simp"], e["trad"]):
            idx.setdefault(key, []).append(e)
    return entries, idx


def _load_junda_freq():
    _ensure_pleco()
    try:
        import openpyxl
    except ImportError:
        return {}
    wb = openpyxl.load_workbook(PLECO_CACHE, data_only=True, read_only=True)
    ws = wb["Junda Frequency"]
    freq = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] is None:
            continue
        simp = str(row[0]).strip()
        freq[simp] = {
            "junda_rank": str(row[1] or "").strip(),
            "junda_count": str(row[2] or "").strip(),
            "junda_pct": str(row[3] or "").strip(),
            "pinyin": str(row[4] or "").strip(),
            "definition": str(row[5] or "").strip(),
        }
    wb.close()
    return freq


def _clean(val):
    v = str(val).strip()
    return "" if v in ("#N/A", "N/A", "None", "") else v


def _merge_meta(entry, pleco_meta, junda_freq):
    simp = entry["simp"]
    pm = pleco_meta.get(simp)
    if pm is None:
        pm = pleco_meta.get(entry.get("trad", ""))
    if pm:
        entry["hsk"] = _clean(pm.get("hsk", ""))
        entry["strokes"] = _clean(pm.get("strokes", ""))
        entry["giga_rank"] = _clean(pm.get("giga_rank", ""))
        entry["junda_rank"] = _clean(pm.get("junda_rank", ""))
    jf = junda_freq.get(simp)
    if jf:
        if not entry.get("pinyin"):
            entry["pinyin"] = _clean(jf.get("pinyin", ""))
        if not entry.get("defs") and jf.get("definition"):
            entry["defs"] = [jf["definition"]]
    entry.setdefault("hsk", "")
    entry.setdefault("strokes", "")
    entry.setdefault("giga_rank", "")
    entry.setdefault("junda_rank", "")
    return entry


# -- Input detection --

def _has_chinese(text):
    return bool(CJK_RANGE.search(text))


# All valid Mandarin pinyin syllables (without tone)
_VALID_PY = {
    "a","ai","an","ang","ao","ba","bai","ban","bang","bao","bei","ben","beng","bi","bian","biao","bie","bin","bing","bo","bu",
    "ca","cai","can","cang","cao","ce","cen","ceng","cha","chai","chan","chang","chao","che","chen","cheng","chi","chong","chou",
    "chu","chua","chuai","chuan","chuang","chui","chun","chuo","ci","cong","cou","cu","cuan","cui","cun","cuo",
    "da","dai","dan","dang","dao","de","dei","den","deng","di","dia","dian","diao","die","ding","diu","dong","dou","du","duan","dui","dun","duo",
    "e","ei","en","eng","er",
    "fa","fan","fang","fei","fen","feng","fiao","fo","fou","fu",
    "ga","gai","gan","gang","gao","ge","gei","gen","geng","gong","gou","gu","gua","guai","guan","guang","gui","gun","guo",
    "ha","hai","han","hang","hao","he","hei","hen","heng","hm","hng","hong","hou","hu","hua","huai","huan","huang","hui","hun","huo",
    "ji","jia","jian","jiang","jiao","jie","jin","jing","jiong","jiu","ju","juan","jue","jun",
    "ka","kai","kan","kang","kao","ke","kei","ken","keng","kong","kou","ku","kua","kuai","kuan","kuang","kui","kun","kuo",
    "la","lai","lan","lang","lao","le","lei","leng","li","lia","lian","liang","liao","lie","lin","ling","liu","lo","long","lou","lu","luan","lun","luo","lv","lue","lü","lve",
    "ma","mai","man","mang","mao","me","mei","men","meng","mi","mian","miao","mie","min","ming","miu","mo","mou","mu",
    "na","nai","nan","nang","nao","ne","nei","nen","neng","ng","ni","nian","niang","niao","nie","nin","ning","niu","nong","nou","nu","nuan","nuo","nv","nve","nü","nve",
    "o","ou",
    "pa","pai","pan","pang","pao","pei","pen","peng","pi","pian","piao","pie","pin","ping","po","pou","pu",
    "qi","qia","qian","qiang","qiao","qie","qin","qing","qiong","qiu","qu","quan","que","qun",
    "ran","rang","rao","re","ren","reng","ri","rong","rou","ru","ruan","rui","run","ruo",
    "sa","sai","san","sang","sao","se","sen","seng","sha","shai","shan","shang","shao","she","shei","shen","sheng","shi","shou","shu","shua","shuai","shuan","shuang","shui","shun","shuo","si","song","sou","su","suan","sui","sun","suo",
    "ta","tai","tan","tang","tao","te","tei","teng","ti","tian","tiao","tie","ting","tong","tou","tu","tuan","tui","tun","tuo",
    "wa","wai","wan","wang","wei","wen","weng","wo","wu",
    "xi","xia","xian","xiang","xiao","xie","xin","xing","xiong","xiu","xu","xuan","xue","xun",
    "ya","yai","yan","yang","yao","ye","yi","yin","ying","yo","yong","you","yu","yuan","yue","yun",
    "za","zai","zan","zang","zao","ze","zei","zen","zeng","zha","zhai","zhan","zhang","zhao","zhe","zhei","zhen","zheng","zhi","zhong","zhou","zhu","zhua","zhuai","zhuan","zhuang","zhui","zhun","zhuo","zi","zong","zou","zu","zuan","zui","zun","zuo",
}
_MAX_PY_LEN = max(len(s) for s in _VALID_PY)


def _segment_pinyin(text):
    """Greedy longest-match segmentation of text into valid pinyin syllables."""
    text = text.lower()
    i = 0
    syllables = []
    while i < len(text):
        matched = False
        for end in range(min(i + _MAX_PY_LEN, len(text)), i, -1):
            if text[i:end] in _VALID_PY:
                syllables.append(text[i:end])
                i = end
                matched = True
                break
        if not matched:
            return None
    return syllables


def _is_pinyin(text):
    text = text.strip().lower()
    if not text or _has_chinese(text):
        return False
    # Allow u: for ü (CEDICT notation) and replace with ü
    text = text.replace("u:", "\u00fc").replace("v", "\u00fc")
    # Tone numbers, ü strongly indicate pinyin
    if re.search(r"[0-9]", text) or "\u00fc" in text:
        return True
    # Must be pure ASCII letters + apostrophe + space
    if not re.match(r"^[a-z' ]+$", text):
        return False
    # Remove apostrophes (used in pinyin like "xi'an"), then segment
    clean = text.replace("'", "")
    parts = clean.split()
    for part in parts:
        if _segment_pinyin(part) is None:
            return False
    return True


def _has_wildcard(text):
    return "*" in text


# -- Search modes --

def _search_chinese(query, entries, index, pleco, junda):
    """Search Chinese text: exact word match + substring matches + char breakdown."""
    result_entries = []
    seen = set()

    def add(e, section):
        key = (e["simp"], e["trad"], e["pinyin"])
        if key not in seen:
            seen.add(key)
            e["_section"] = section
            _merge_meta(e, pleco, junda)
            result_entries.append(e)

    # 1. Exact whole-word match
    exact = index.get(query, [])
    for e in exact:
        add(e, "exact")

    # 2. Substring matches (entries containing the query)
    for e in entries:
        if query in e["simp"] or query in e["trad"]:
            add(e, "substring")

    # 3. Character breakdown for multi-char queries
    chars = [c for c in query if CJK_RANGE.match(c)]
    char_results = []
    if len(chars) > 1:
        for ch in chars:
            ch_entries = [e for e in entries if e["simp"] == ch or e["trad"] == ch]
            if ch_entries:
                merged = []
                for e in ch_entries:
                    e_copy = dict(e)
                    _merge_meta(e_copy, pleco, junda)
                    e_copy["_section"] = f"char:{ch}"
                    merged.append(e_copy)
                char_results.append({"char": ch, "entries": merged})
                result_entries.extend(merged)

    return {
        "query": query,
        "mode": "chinese",
        "entries": result_entries,
        "characters": char_results,
        "total": len(result_entries),
    }


def _search_pinyin(query, entries, pleco, junda):
    """Search by pinyin (with/without tone numbers, multi-syllable ok)."""
    nq = _norm_pinyin(query).replace(" ", "")
    result = []
    seen = set()
    for e in entries:
        for py_part in e["pinyin"].split("/"):
            np = _norm_pinyin(py_part).replace(" ", "")
            if nq == np:
                key = (e["simp"], e["trad"], e["pinyin"])
                if key not in seen:
                    seen.add(key)
                    e["_section"] = "pinyin"
                    _merge_meta(e, pleco, junda)
                    result.append(e)
                break
    return {
        "query": query,
        "mode": "pinyin",
        "entries": result,
        "characters": [],
        "total": len(result),
    }


def _search_english(query, entries, pleco, junda):
    """Search English definitions."""
    q = query.lower().strip()
    result = []
    seen = set()
    # Support wildcard in English
    if _has_wildcard(q):
        pat = fnmatch.translate(q.lower())
        regex = re.compile(pat)
        for e in entries:
            for d in e["defs"]:
                if regex.search(d.lower()):
                    key = (e["simp"], e["trad"], e["pinyin"])
                    if key not in seen:
                        seen.add(key)
                        e["_section"] = "english"
                        _merge_meta(e, pleco, junda)
                        result.append(e)
                    break
    else:
        for e in entries:
            for d in e["defs"]:
                if q in d.lower():
                    key = (e["simp"], e["trad"], e["pinyin"])
                    if key not in seen:
                        seen.add(key)
                        e["_section"] = "english"
                        _merge_meta(e, pleco, junda)
                        result.append(e)
                    break
    return {
        "query": query,
        "mode": "english",
        "entries": result,
        "characters": [],
        "total": len(result),
    }


def _search_wildcard_chinese(query, entries, pleco, junda):
    """Wildcard search in Chinese simplified/traditional."""
    pat = fnmatch.translate(query)
    regex = re.compile(pat)
    result = []
    seen = set()
    for e in entries:
        if regex.search(e["simp"]) or regex.search(e["trad"]):
            key = (e["simp"], e["trad"], e["pinyin"])
            if key not in seen:
                seen.add(key)
                e["_section"] = "wildcard"
                _merge_meta(e, pleco, junda)
                result.append(e)
    return {
        "query": query,
        "mode": "chinese",
        "entries": result,
        "characters": [],
        "total": len(result),
    }


_cedict_entries = None
_cedict_index = None
_pleco_meta = None
_junda_freq = None

_wikdict_pl_entries = None
_wikdict_pl_index = None

_ccpl_entries = None
_ccpl_index = None


def _ensure_loaded():
    global _cedict_entries, _cedict_index, _pleco_meta, _junda_freq
    if _cedict_index is not None:
        return
    _cedict_entries = load_entries()
    _cedict_index = build_index(_cedict_entries)
    _pleco_meta = _load_pleco_export()
    _junda_freq = _load_junda_freq()


def _ensure_pl_loaded():
    global _wikdict_pl_entries, _wikdict_pl_index
    if _wikdict_pl_index is not None:
        return
    _wikdict_pl_entries, _wikdict_pl_index = _load_wikdict_pl()


def _ensure_ccpl_loaded():
    global _ccpl_entries, _ccpl_index
    if _ccpl_index is not None:
        return
    _ensure_loaded()
    _ensure_pl_loaded()
    _ccpl_entries = []
    for e in _cedict_entries:
        m = dict(e)
        m["defs_en"] = list(e["defs"])
        pl_entries = _wikdict_pl_index.get(e["simp"], [])
        if not pl_entries:
            pl_entries = _wikdict_pl_index.get(e["trad"], [])
        if pl_entries:
            polish = set()
            for pe in pl_entries:
                for d in pe["defs"]:
                    polish.add(d)
            m["defs"] = sorted(polish)
        else:
            m["defs"] = list(e["defs"])
        _ccpl_entries.append(m)
    _ccpl_index = build_index(_ccpl_entries)


def _enrich_pl_entry(entry, cedict_index, pleco_meta):
    """Add pinyin, HSK, strokes to a WikDict entry from CC-CEDICT/Pleco."""
    simp = entry["simp"]
    ce_list = cedict_index.get(simp, [])
    if ce_list:
        entry["pinyin"] = ce_list[0].get("pinyin", "")
    entry.setdefault("pinyin", "")
    pm = pleco_meta.get(simp)
    if pm:
        entry["hsk"] = _clean(pm.get("hsk", ""))
        entry["strokes"] = _clean(pm.get("strokes", ""))
    entry.setdefault("hsk", "")
    entry.setdefault("strokes", "")
    return entry


def _search_polish(query, entries, index, cedict_index, pleco_meta):
    """Search Chinese-Polish WikDict: exact match + substring + Polish text."""
    result_entries = []
    seen = set()

    def add(e, section):
        key = e["simp"]
        if key not in seen:
            seen.add(key)
            e["_section"] = section
            _enrich_pl_entry(e, cedict_index, pleco_meta)
            result_entries.append(e)

    # 1. Exact match
    exact = index.get(query, [])
    for e in exact:
        add(e, "exact")

    # 2. Substring (Chinese)
    for e in entries:
        if query in e["simp"]:
            add(e, "substring")

    # 3. Polish definition search
    q = query.lower().strip()
    for e in entries:
        for d in e["defs"]:
            if q in d.lower():
                add(e, "wyszukiwanie PL")
                break

    # 4. Character breakdown for multi-char queries
    chars = [c for c in query if CJK_RANGE.match(c)]
    char_results = []
    if len(chars) > 1:
        for ch in chars:
            ch_entries = index.get(ch, [])
            if ch_entries:
                merged = [dict(e) for e in ch_entries]
                for e in merged:
                    e["_section"] = f"znak:{ch}"
                    _enrich_pl_entry(e, cedict_index, pleco_meta)
                char_results.append({"char": ch, "entries": merged})
                result_entries.extend(merged)

    return {
        "query": query,
        "mode": "polish",
        "entries": result_entries,
        "characters": char_results,
        "total": len(result_entries),
    }


def _search_pl_pinyin(query, wikdict_index, cedict_index, pleco_meta):
    """Pinyin search in Polish dictionary: find via CC-CEDICT, get Polish defs from WikDict."""
    _ensure_loaded()
    nq = _norm_pinyin(query).replace(" ", "")
    result = []
    seen = set()
    for e in _cedict_entries:
        for py_part in e["pinyin"].split("/"):
            np = _norm_pinyin(py_part).replace(" ", "")
            if nq == np:
                key = e["simp"]
                if key not in seen:
                    seen.add(key)
                    pl_entries = wikdict_index.get(key, [])
                    if pl_entries:
                        pe = dict(pl_entries[0])
                        pe["pinyin"] = e["pinyin"]
                        pe["_section"] = "pinyin"
                        _enrich_pl_entry(pe, cedict_index, pleco_meta)
                        result.append(pe)
                    else:
                        pm = pleco_meta.get(key, {})
                        result.append({
                            "trad": e["trad"], "simp": e["simp"],
                            "pinyin": e["pinyin"], "defs": [],
                            "hsk": _clean(pm.get("hsk", "")),
                            "strokes": _clean(pm.get("strokes", "")),
                            "_section": "pinyin",
                        })
                break
    return {
        "query": query, "mode": "polish",
        "entries": result, "characters": [], "total": len(result),
    }


def lookup(text, lang="en"):
    """Search dictionary with different backends.

    lang="en": CC-CEDICT Chinese-English (auto-detects Chinese/Pinyin/English, wildcards, character breakdown)
    lang="pl": WikDict Chinese-Polish (Chinese exact/substring, Polish definition search, pinyin via CC-CEDICT)
    lang="cc-pl": CC-CEDICT merged with Polish definitions (CC-CEDICT structure + WikDict Polish, fallback to English)
    """
    q = text.strip()
    if not q:
        return {"query": "", "mode": "unknown", "entries": [], "characters": [], "total": 0}

    if lang == "pl":
        _ensure_pl_loaded()
        _ensure_loaded()
        if _has_chinese(q):
            return _search_polish(q, _wikdict_pl_entries, _wikdict_pl_index, _cedict_index, _pleco_meta)
        if _is_pinyin(q):
            return _search_pl_pinyin(q, _wikdict_pl_index, _cedict_index, _pleco_meta)
        return _search_polish(q, _wikdict_pl_entries, _wikdict_pl_index, _cedict_index, _pleco_meta)

    if lang == "cc-pl":
        _ensure_ccpl_loaded()
        if _has_chinese(q):
            if _has_wildcard(q):
                return _search_wildcard_chinese(q, _ccpl_entries, _pleco_meta, _junda_freq)
            return _search_chinese(q, _ccpl_entries, _ccpl_index, _pleco_meta, _junda_freq)
        if _is_pinyin(q):
            if _has_wildcard(q):
                return _search_english(q, _ccpl_entries, _pleco_meta, _junda_freq)
            return _search_pinyin(q, _ccpl_entries, _pleco_meta, _junda_freq)
        return _search_english(q, _ccpl_entries, _pleco_meta, _junda_freq)

    _ensure_loaded()
    if _has_chinese(q):
        if _has_wildcard(q):
            return _search_wildcard_chinese(q, _cedict_entries, _pleco_meta, _junda_freq)
        return _search_chinese(q, _cedict_entries, _cedict_index, _pleco_meta, _junda_freq)
    if _is_pinyin(q):
        if _has_wildcard(q):
            return _search_english(q, _cedict_entries, _pleco_meta, _junda_freq)
        return _search_pinyin(q, _cedict_entries, _pleco_meta, _junda_freq)
    return _search_english(q, _cedict_entries, _pleco_meta, _junda_freq)


def lookup_chinese_simple(text):
    """Simple exact-match lookup (original behavior), returns list of entries."""
    _ensure_loaded()
    key = text.strip()
    results = _cedict_index.get(key, [])
    if results:
        for r in results:
            _merge_meta(r, _pleco_meta, _junda_freq)
    else:
        pm = _pleco_meta.get(key)
        jf = _junda_freq.get(key)
        if pm or jf:
            entry = {
                "trad": pm["trad"] if pm else key,
                "simp": key,
                "pinyin": (pm or {}).get("pinyin", "") or (jf or {}).get("pinyin", ""),
                "defs": [jf["definition"]] if jf and jf.get("definition") else [],
            }
            _merge_meta(entry, _pleco_meta, _junda_freq)
            results = [entry]
    return results
